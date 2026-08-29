"""Safe, format-aware portfolio statement ingestion.

CSV remains the canonical parser: XLSX worksheets and PDF tables are converted
to in-memory rows and then sent through :mod:`ystocker.portfolio_csv`.  Keeping
one column-mapping and validation path matters because every import format gets
the same preview-before-save contract and the same protections against totals,
prose, malformed numbers and invented symbols.

Uploaded document contents are data only.  Nothing in a workbook cell or PDF
page is interpreted as an instruction or executed.
"""

from __future__ import annotations

import csv
import io
import logging
import re
from pathlib import Path
from typing import Any, Iterable

from ystocker import portfolio_csv

log = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = frozenset({".csv", ".txt", ".xlsx", ".pdf"})
MAX_PDF_PAGES = 50
MAX_TABLES = 100
MAX_SHEET_ROWS = portfolio_csv.MAX_ROWS + portfolio_csv.MAX_HEADER_SCAN


def parse(raw: bytes, *, filename: str = "", content_type: str = "") -> portfolio_csv.ParseResult:
    """Parse a CSV, XLSX workbook or PDF statement without writing it to disk."""
    suffix = Path(filename or "").suffix.lower()
    mime = (content_type or "").split(";", 1)[0].strip().lower()

    if suffix and suffix not in SUPPORTED_EXTENSIONS:
        result = portfolio_csv.ParseResult()
        result.error = (f"Unsupported file type '{suffix}'. Upload a CSV, TXT, "
                        "XLSX or PDF file.")
        return result

    if suffix == ".pdf" or mime == "application/pdf" or raw.startswith(b"%PDF-"):
        return _parse_pdf(raw)
    if (suffix == ".xlsx"
            or mime == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"):
        return _parse_xlsx(raw)
    return portfolio_csv.parse(raw)


def _csv_text(rows: Iterable[Iterable[Any]]) -> str:
    stream = io.StringIO(newline="")
    writer = csv.writer(stream)
    for row in rows:
        writer.writerow([_cell_text(cell) for cell in row])
    return stream.getvalue()


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    # PDF extractors commonly preserve line wraps inside one table cell.  They
    # are layout artifacts, not delimiters, and break heading matching.
    return re.sub(r"\s+", " ", str(value)).strip()[:500]


def _parse_xlsx(raw: bytes) -> portfolio_csv.ParseResult:
    result = portfolio_csv.ParseResult(encoding="XLSX")
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException

        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True,
                                 keep_links=False)
    except ImportError:
        result.error = "XLSX import is temporarily unavailable on this server"
        return result
    except (InvalidFileException, OSError, ValueError, KeyError, EOFError) as exc:
        result.error = f"Could not read the XLSX workbook: {exc}"
        return result
    except Exception as exc:  # BadZipFile and XML parser failures vary by version.
        log.info("portfolio_import: invalid XLSX: %s", exc)
        result.error = "Could not read the XLSX workbook; it may be damaged or encrypted"
        return result

    try:
        visible = [sheet for sheet in workbook.worksheets
                   if getattr(sheet, "sheet_state", "visible") == "visible"]
        sheets = visible or list(workbook.worksheets)
        parsed: list[tuple[str, portfolio_csv.ParseResult]] = []
        failures: list[str] = []
        for sheet in sheets:
            rows = []
            truncated = False
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                if index >= MAX_SHEET_ROWS:
                    truncated = True
                    break
                rows.append(row)
            if not rows or all(all(cell is None or not str(cell).strip() for cell in row)
                               for row in rows):
                continue
            sheet_result = portfolio_csv.parse(_csv_text(rows))
            if sheet_result.ok:
                if truncated:
                    sheet_result.warnings.append(
                        f"Sheet '{sheet.title}' was limited to {MAX_SHEET_ROWS} rows")
                for position in sheet_result.rows:
                    if not position.account:
                        position.account = str(sheet.title)[:60]
                parsed.append((str(sheet.title), sheet_result))
            else:
                failures.append(f"{sheet.title}: {sheet_result.error}")
        return _combine(parsed, format_name="Excel workbook",
                        encoding=f"XLSX · {len(sheets)} sheet(s)", failures=failures)
    finally:
        workbook.close()


def _parse_pdf(raw: bytes) -> portfolio_csv.ParseResult:
    result = portfolio_csv.ParseResult(encoding="PDF")
    if not raw.startswith(b"%PDF-"):
        result.error = "The uploaded file is not a valid PDF"
        return result
    try:
        import pdfplumber
    except ImportError:
        result.error = "PDF import is temporarily unavailable on this server"
        return result

    parsed: list[tuple[str, portfolio_csv.ParseResult]] = []
    failures: list[str] = []
    try:
        with pdfplumber.open(io.BytesIO(raw)) as document:
            if len(document.pages) > MAX_PDF_PAGES:
                result.error = (f"PDF has {len(document.pages)} pages; the limit is "
                                f"{MAX_PDF_PAGES}")
                return result

            table_count = 0
            for page_number, page in enumerate(document.pages, start=1):
                page_parsed = False
                # Borderless brokerage statements need text-aligned table
                # discovery. Try it when line discovery found no *usable* table,
                # not merely when it found no rectangles at all.
                table_sets = (page.extract_tables(), None)
                for table_set_index in range(2):
                    tables = table_sets[table_set_index]
                    if table_set_index == 1:
                        tables = page.extract_tables({
                            "vertical_strategy": "text",
                            "horizontal_strategy": "text",
                            "min_words_vertical": 2,
                            "min_words_horizontal": 1,
                        })
                    for table in tables or []:
                        table_count += 1
                        if table_count > MAX_TABLES:
                            result.error = f"PDF contains more than {MAX_TABLES} tables"
                            return result
                        table_result = portfolio_csv.parse(
                            _csv_text(table[:MAX_SHEET_ROWS]))
                        if table_result.ok:
                            parsed.append((f"page {page_number}", table_result))
                            page_parsed = True
                    if page_parsed:
                        break

                if not page_parsed:
                    text_rows = _pdf_text_rows(page.extract_text() or "")
                    if text_rows:
                        text_result = portfolio_csv.parse(_csv_text(text_rows))
                        if text_result.ok:
                            parsed.append((f"page {page_number}", text_result))
                            page_parsed = True
                        elif text_result.error:
                            failures.append(f"page {page_number}: {text_result.error}")

            return _combine(parsed, format_name="PDF statement",
                            encoding=f"PDF · {len(document.pages)} page(s)",
                            failures=failures)
    except Exception as exc:
        log.info("portfolio_import: unreadable PDF: %s", exc)
        result.error = "Could not read the PDF; it may be damaged, encrypted or image-only"
        return result


def _pdf_text_rows(text: str) -> list[list[str]]:
    """Best-effort rows for PDFs whose columns are aligned with whitespace."""
    rows: list[list[str]] = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        cells = re.split(r"\t+|\s{2,}|\s*\|\s*", line)
        if len(cells) > 1:
            rows.append(cells)
    return rows


def _combine(parts: list[tuple[str, portfolio_csv.ParseResult]], *,
             format_name: str, encoding: str,
             failures: list[str]) -> portfolio_csv.ParseResult:
    """Combine independently parsed sheets/tables into one preview result."""
    combined = portfolio_csv.ParseResult(broker=format_name, encoding=encoding)
    if not parts:
        detail = failures[0] if failures else "no readable position table was found"
        combined.error = f"No positions found in {format_name}. {detail}"
        return combined

    headings: dict[str, list[str]] = {}
    for source, part in parts:
        remaining = portfolio_csv.MAX_ROWS - len(combined.rows)
        if remaining <= 0:
            combined.warnings.append(
                f"Stopped at {portfolio_csv.MAX_ROWS} positions; the rest were not read")
            break
        combined.rows.extend(part.rows[:remaining])
        combined.skipped.extend(part.skipped[:portfolio_csv.MAX_ROWS
                                             - len(combined.skipped)])
        if len(part.rows) > remaining:
            combined.warnings.append(
                f"Stopped at {portfolio_csv.MAX_ROWS} positions; the rest were not read")
        combined.warnings.extend(part.warnings)
        for field, heading in part.mapping.items():
            headings.setdefault(field, [])
            if heading not in headings[field]:
                headings[field].append(heading)
        if part.header_line and not combined.header_line:
            combined.header_line = part.header_line
        if part.broker and part.broker != format_name:
            combined.warnings.append(f"{source}: detected {part.broker}")

    combined.mapping = {field: " / ".join(values)
                        for field, values in headings.items()}
    if failures:
        combined.warnings.append(
            f"{len(failures)} sheet/page table(s) contained no importable positions")

    counts: dict[str, int] = {}
    for row in combined.rows:
        counts[row.symbol] = counts.get(row.symbol, 0) + 1
    duplicates = sorted(symbol for symbol, count in counts.items() if count > 1)
    if duplicates and not any("appears more than once" in warning
                              for warning in combined.warnings):
        combined.warnings.append(
            "appears more than once (separate accounts?): "
            + ", ".join(duplicates[:10]))
    return combined
